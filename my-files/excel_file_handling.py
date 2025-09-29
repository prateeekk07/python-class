import openpyxl

# wb = openpyxl.Workbook()
# sheet = wb.active

#1st Way

# sheet["A1"] = "Hello"
# sheet["B1"] = "World!"
# sheet["D5"] = "dUMMY data"

# #2nd Way

# sheet.cell(row=10, column=6).value = "Python"

# for i in range(1,11):
#     sheet.cell(row=i, column=i).value = i

FILE_PATH = r"D:\python b14\python-class\my-files\test.xlsx"

wb = openpyxl.load_workbook(FILE_PATH)
print(wb.__dir__)

# wb.save(filename=r"D:\python b14\python-class\my-files\test.xlsx")
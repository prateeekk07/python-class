# for data store  -->
# library --> openpyxl --> 

import openpyxl

# wb = openpyxl.Workbook() # new wb
# sheet = wb.active

# 2nd way
# sheet.cell(row=10, column=6).value = "Python"
# sheet.cell(row=10, column=8).value = "Java"

# 1st way
# sheet["D1"] = "Hello"
# sheet["F1"] = "World!"
# sheet["Z50"] = "Dummy Data"
# sheet["H24"] = 5565.545


# for i in range(1, 11):
    # sheet.cell(row=i, column=i).value = i


# for i in range(1, 11): # 3
#     for j in range(1, 11):
#         sheet.cell(row=i, column=j).value = chr(64+j)

FILE_PATH = r"E:\Python-B14\Files\file_handling\employee.xlsx"

wb = openpyxl.load_workbook(FILE_PATH)
sheet = wb["Data"]
# print(sheet[5])
# print(sheet.max_column)
# print(sheet.max_row)

# print(dir(sheet))

# for row in sheet.iter_rows(min_row=3, max_col=5):
#     for d in row:
#         print(d.value, end="  ")
#     print()

# for row in sheet["A1:F9"]:
#     for d in row:
#         print(d.value, end=" ")
#     print()

# for i in range(1, sheet.max_row+1): # 2
#     for j in range(1, sheet.max_column+1):
#         print(sheet.cell(row=i, column=j).value, end=" ")
#     print()


# sheet = wb["Sheet3"]
# print(sheet["A1"].value)
# print(wb.sheetnames)
# wb.active

# wb.save(filename=)

# generate
# employee data insert --> class dict --> excel
# employee data read --> excel --> class dict

# class --> objects



